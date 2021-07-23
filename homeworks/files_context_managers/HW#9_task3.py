"""
Task 3
Використовуючи openpyxl (або будь-яку іншу зручну для вас бібліотеку),
напишіть контекстний менеджер для роботи з ексель.
Даний менеджер повинен бути аналогом методу open()
"""

import openpyxl


class OpenExel:
    def __init__(self, path):
        self.file = openpyxl.load_workbook(path)
        self.sheet = self.file.active
        self.rows = self.sheet.max_row
        self.cols = self.sheet.max_column

    def __enter__(self):
        return self

    def __str__(self):
        for i in range(1, self.rows + 1):
            string = ''
            for j in range(1, self.cols + 1):
                cell = self.sheet.cell(row=i, column=j)
                string = string + str(cell.value) + ' '
            print(string)
        print(self.sheet)
        return f''

    def __exit__(self, exc_type, exc_val, exc_tb):
        if exc_type is None:
            self.file.close()
        elif exc_type is Exception:
            print('Exception')
            return True


with OpenExel('exam.xlsx') as file_exel:
    print(file_exel)
    raise Exception


# № ПІБ Гуртожиток
# 1 Марценюк Олексій 1 гуртожиток
# 2 Колотило Сергій Андрійович 1 гуртожиток
# 3 Костіна Аліна Олегівна 3 гуртожиток
# 4 Карпанець Олександр Сергійович 3 гуртожиток
# 5 Крикля Владислав Андрійович 3 гуртожиток
#
# Exception
#


import pandas

class OpenExelPandas:
    def __init__(self, path, sheet):
        self.excel_data = pandas.read_excel(path, sheet_name=sheet)

    def __enter__(self):
        return self.excel_data

    def __exit__(self, exc_type, exc_val, exc_tb):
        if exc_type is None:
            self.excel_data.close()
        elif exc_type is Exception:
            print('Exception')
            return True

with OpenExelPandas('exam.xlsx','Sheet1') as file_exel:
    print(file_exel)
    raise Exception


#        №                               ПІБ     Гуртожиток
# 0      1                  Марценюк Олексій   1 гуртожиток
# 1      2        Колотило Сергій Андрійович   1 гуртожиток
# 2      3            Костіна Аліна Олегівна   3 гуртожиток
# 3      4    Карпанець Олександр Сергійович   3 гуртожиток
# 4      5       Крикля Владислав Андрійович   3 гуртожиток
#
# [5 rows x 3 columns]
# Exception
